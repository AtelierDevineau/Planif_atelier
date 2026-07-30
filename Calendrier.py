import streamlit as st
import streamlit.components.v1 as components
from streamlit_calendar import calendar
from datetime import date, timedelta
from donnees import Options_cal, build_absences_cal, sauvegarder_ressources_github
from donnees import COULEURS_TACHES, ORDRE_TACHES, get_couleur_tache, fmt_date
import pandas as pd
import io

# -------------------------------------------------------
# FONCTIONS D'EXPORT
# -------------------------------------------------------

def _build_export_df(projets_data, data_proj):
    """Construit un DataFrame avec toutes les sous-tâches pour export."""
    rows = []
    for projet in projets_data:
        nom_projet = projet["projet"]
        for sous_tache in projet["sous_taches"]:
            nom_tache = sous_tache["tache"]
            noms = [a["Nom"] for a in data_proj.get(nom_projet, {}).get(nom_tache, {}).get("Assignations", [])]
            rows.append({
                "Projet": nom_projet,
                "Sous-tâche": nom_tache,
                "Début": sous_tache["start"],
                "Fin": sous_tache["end"],
                "Ressources assignées": ", ".join(noms)
            })
    return pd.DataFrame(rows)

def _to_excel_gantt(projets_data, data_proj):
    """
    Génère un fichier Excel avec un Gantt coloré par cellules.
    - Colonnes = semaines (en-tête lundi JJ/MM) subdivisées en jours
    - Lignes = sous-tâches (2 lignes : nom projet + ressources)
    - Cellules colorées selon le type de sous-tâche
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    # Calculer la plage de dates
    toutes_dates = [
        date.fromisoformat(st_data[cle])
        for p in projets_data
        for st_data in p.get("sous_taches", [])
        for cle in ("start", "end")
    ]
    if not toutes_dates:
        return io.BytesIO().getvalue()

    today = date.today()
    date_debut = min(min(toutes_dates), today) - timedelta(days=7)
    date_fin = max(toutes_dates) + timedelta(days=7)
    jours = []
    cur = date_debut
    while cur <= date_fin:
        jours.append(cur)
        cur += timedelta(days=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt"

    # ── En-tête : une colonne label + une colonne par jour ──────────────────
    LABEL_COL = 1
    JOUR_START_COL = 2

    # Colonne label
    ws.column_dimensions[get_column_letter(LABEL_COL)].width = 22

    # En-têtes semaines (ligne 1) et jours (ligne 2)
    lundi_courant = None
    for idx, jour in enumerate(jours):
        col = JOUR_START_COL + idx
        ws.column_dimensions[get_column_letter(col)].width = 3.5

        # Ligne 1 : label semaine sur le lundi
        if jour.weekday() == 0:
            lundi_courant = col
            cell_sem = ws.cell(row=1, column=col)
            cell_sem.value = f"Lun. {jour.day}/{jour.month}"
            cell_sem.font = Font(bold=True, size=8)
            cell_sem.alignment = Alignment(horizontal="left")
            cell_sem.fill = PatternFill("solid", fgColor="F5F5F5")

        # Ligne 2 : numéro du jour
        cell_j = ws.cell(row=2, column=col)
        cell_j.value = jour.day
        cell_j.font = Font(size=7)
        cell_j.alignment = Alignment(horizontal="center")
        if jour == today:
            cell_j.fill = PatternFill("solid", fgColor="FFFACD")

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 14

    # ── Lignes de données ────────────────────────────────────────────────────
    row = 3
    for type_tache in ORDRE_TACHES:
        couleur_hex = COULEURS_TACHES.get(type_tache, "#A0A0A0").lstrip("#")
        couleur_sub = couleur_hex + "99"  # opacité simulée en Excel n'existe pas,
        # on utilise une version plus claire
        r = int(couleur_hex[0:2], 16)
        g = int(couleur_hex[2:4], 16)
        b = int(couleur_hex[4:6], 16)
        # Version claire pour la ligne ressources (mélange avec blanc)
        r2 = min(255, r + (255 - r) // 2)
        g2 = min(255, g + (255 - g) // 2)
        b2 = min(255, b + (255 - b) // 2)
        couleur_claire = f"{r2:02X}{g2:02X}{b2:02X}"

        for projet in projets_data:
            nom_projet = projet["projet"]
            for sous_tache in projet["sous_taches"]:
                nom_tache = sous_tache["tache"]
                if (nom_tache if nom_tache in ORDRE_TACHES else "Autre") != type_tache:
                    continue

                debut_st = date.fromisoformat(sous_tache["start"])
                fin_st = date.fromisoformat(sous_tache["end"])
                noms = [a["Nom"] for a in data_proj.get(nom_projet, {}).get(nom_tache, {}).get("Assignations", [])]
                noms_str = ", ".join(noms)

                # Ligne 1 : nom projet
                ws.cell(row=row, column=LABEL_COL).value = nom_projet
                ws.cell(row=row, column=LABEL_COL).font = Font(bold=True, size=8)
                ws.row_dimensions[row].height = 16

                # Ligne 2 : ressources
                ws.cell(row=row+1, column=LABEL_COL).value = noms_str
                ws.cell(row=row+1, column=LABEL_COL).font = Font(size=7, italic=True)
                ws.row_dimensions[row+1].height = 13

                # Colorier les cellules des jours concernés
                for idx, jour in enumerate(jours):
                    col = JOUR_START_COL + idx
                    if debut_st <= jour < fin_st:
                        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=couleur_hex)
                        ws.cell(row=row+1, column=col).fill = PatternFill("solid", fgColor=couleur_claire)
                    elif jour == today:
                        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFFACD")
                        ws.cell(row=row+1, column=col).fill = PatternFill("solid", fgColor="FFFACD")

                row += 2

        # Ligne de séparation entre types de tâches
        if any(
            (st["tache"] if st["tache"] in ORDRE_TACHES else "Autre") == type_tache
            for p in projets_data for st in p.get("sous_taches", [])
        ):
            for col_idx in range(len(jours) + 1):
                ws.cell(row=row, column=LABEL_COL + col_idx).border = Border(
                    bottom=Side(style="thin", color="CCCCCC")
                )
            row += 1

    # Figer la première colonne et les deux premières lignes
    ws.freeze_panes = "B3"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# -------------------------------------------------------
# UTILITAIRES
# -------------------------------------------------------

def get_lundis(date_debut, date_fin):
    lundis = []
    cur = date_debut - timedelta(days=date_debut.weekday())
    while cur <= date_fin:
        lundis.append(cur)
        cur += timedelta(weeks=1)
    return lundis

def get_jours(date_debut, date_fin):
    jours = []
    cur = date_debut
    while cur <= date_fin:
        jours.append(cur)
        cur += timedelta(days=1)
    return jours

def get_noms_assignes(nom_projet, nom_tache, data_proj):
    data_st = data_proj.get(nom_projet, {}).get(nom_tache, {})
    return [a["Nom"] for a in data_st.get("Assignations", [])]


# -------------------------------------------------------
# CONSTRUCTION DU GANTT TABLEAU
# -------------------------------------------------------

def gantt_tableau(projets_data, data_proj, date_debut_grille, date_fin_grille, gantt_id="gantt", font_face=""):
    today = date.today()

    jours = get_jours(date_debut_grille, date_fin_grille)
    lundis = get_lundis(date_debut_grille, date_fin_grille)
    nb_jours = len(jours)
    jour_index = {j: i for i, j in enumerate(jours)}

    # Colonnes avec séparateurs de semaine
    colonnes = []
    for jour in jours:
        if jour.weekday() == 0 and jour != jours[0]:
            colonnes.append({"type": "sep"})
        colonnes.append({"type": "jour", "jour": jour})

    col_index = {}
    for i, col in enumerate(colonnes):
        if col["type"] == "jour":
            col_index[col["jour"]] = i

    nb_cols = len(colonnes)
    nb_seps = sum(1 for c in colonnes if c["type"] == "sep")
    largeur_totale = nb_jours * 14 + nb_seps * 3

    # Regrouper toutes les sous-tâches par type
    # Structure : {type_tache: [(nom_projet, sous_tache_data), ...]}
    taches_par_type = {t: [] for t in ORDRE_TACHES}
    for projet in projets_data:
        nom_projet = projet["projet"]
        for sous_tache in projet["sous_taches"]:
            nom_tache = sous_tache["tache"]
            type_tache = nom_tache if nom_tache in ORDRE_TACHES else "Autre"
            taches_par_type[type_tache].append((nom_projet, sous_tache))

    # Scroll vers aujourd'hui
    if today >= date_debut_grille:
        jours_avant = (today - date_debut_grille).days
        seps_avant = sum(1 for j in jours if j <= today and j.weekday() == 0 and j != date_debut_grille)
        scroll_px = max(0, jours_avant * 14 + seps_avant * 3)
    else:
        scroll_px = 0

    css = f"""
    <style>
    {font_face}
    body, table, td, th {{
        font-family: 'GTWalsheim', sans-serif;
    }}
    .gantt-wrap {{ overflow-x: auto; width: 100%; }}
    .gantt-table {{
        border-collapse: separate;
        border-spacing: 0;
        width: {largeur_totale}px;
        table-layout: fixed;
        font-size: 0.82em;
    }}
    .gh {{
        text-align: left;
        padding: 4px 6px;
        font-weight: bold;
        font-size: 0.85em;
        background: #f5f5f5;
        border-bottom: 2px solid #ccc;
        border-top: none;
        border-right: none;
        border-left: none;
        white-space: nowrap;
        overflow: hidden;
    }}
    .gh-sep {{
        background: #999;
        width: 3px;
        padding: 0;
        border: none;
    }}
    .gc {{
        padding: 0;
        height: 24px;
        border-right: 1px solid #f0f0f0;
        border-top: none;
        border-bottom: none;
        border-left: none;
    }}
    .gc-sep {{
        background: #bbb;
        width: 3px;
        padding: 0;
        border: none;
        height: 24px;
    }}
    .gc-sep2 {{
        background: #bbb;
        width: 3px;
        padding: 0;
        border: none;
        height: 20px;
    }}
    .gb1 {{
        padding: 0 8px;
        height: 24px;
        vertical-align: middle;
        text-align: center;
        font-weight: bold;
        font-size: 0.85em;
        color: #fff;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
        border: none;
    }}
    .gb1-dark {{
        padding: 0 8px;
        height: 24px;
        vertical-align: middle;
        text-align: center;
        font-weight: bold;
        font-size: 0.85em;
        color: #222;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
        border: none;
    }}
    .gb2 {{
        padding: 0 8px;
        height: 20px;
        vertical-align: middle;
        font-size: 0.78em;
        color: #333;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
        border: none;
    }}
    .gsep td {{
        border-bottom: 2px solid #ccc !important;
    }}
    .gc-today {{
        padding: 0;
        height: 24px;
        border-right: 1px solid #f0f0f0;
        border-top: none;
        border-bottom: none;
        border-left: none;
        background-color: #FFFACD;
    }}
    .gc2-today {{
        padding: 0;
        height: 20px;
        border-right: 1px solid #f0f0f0;
        border-top: none;
        border-bottom: none;
        border-left: none;
        background-color: #FFFACD;
    }}
    .gc2 {{
        padding: 0;
        height: 20px;
        border-right: 1px solid #f0f0f0;
        border-top: none;
        border-bottom: none;
        border-left: none;
    }}
    </style>
    """

    # ── HEADER ───────────────────────────────────────────────────────────────
    header = '<tr>'
    for lundi in lundis:
        jours_sem = [j for j in jours if lundi <= j < lundi + timedelta(weeks=1)]
        if not jours_sem:
            continue
        if lundi != lundis[0]:
            header += '<th class="gh-sep"></th>'
        label = f"Lun. {lundi.day}/{lundi.month}"
        header += f'<th colspan="{len(jours_sem)}" class="gh">{label}</th>'
    header += '</tr>'

    # ── COLGROUP ─────────────────────────────────────────────────────────────
    colgroup = '<colgroup>'
    for col in colonnes:
        if col["type"] == "sep":
            colgroup += '<col style="width:3px;min-width:3px;max-width:3px;">'
        else:
            colgroup += '<col style="width:14px;min-width:14px;">'
    colgroup += '</colgroup>'

    # ── LIGNES ───────────────────────────────────────────────────────────────
    rows = ""

    for type_idx, type_tache in enumerate(ORDRE_TACHES):
        taches = taches_par_type.get(type_tache, [])
        if not taches:
            continue

        couleur = COULEURS_TACHES.get(type_tache, "#A0A0A0")
        # Texte blanc sur couleurs sombres, noir sur claires
        couleurs_claires = {"#F1C84E", "#7EB8F7", "#A0C45A"}
        cls_texte = "gb1-dark" if couleur in couleurs_claires else "gb1"
        couleur_sub = couleur + "99"

        for st_idx, (nom_projet, sous_tache) in enumerate(taches):
            nom_tache = sous_tache["tache"]
            debut_st = date.fromisoformat(sous_tache["start"])
            fin_st = date.fromisoformat(sous_tache["end"])
            noms = get_noms_assignes(nom_projet, nom_tache, data_proj)
            noms_str = ", ".join(noms)

            is_last = st_idx == len(taches) - 1

            debut_visible = max(debut_st, date_debut_grille)
            fin_visible = min(fin_st, date_fin_grille)
            dans_fenetre = debut_visible < fin_visible

            if dans_fenetre:
                ci_debut = col_index.get(debut_visible, 0)
                ci_fin = col_index.get(fin_visible, nb_cols)
                colspan_barre = ci_fin - ci_debut
            else:
                ci_debut = None
                colspan_barre = 0

            # Ligne 1 : nom du projet dans la barre — cliquable
            row1 = '<tr>'
            ci = 0
            while ci < nb_cols:
                col = colonnes[ci]
                if col["type"] == "sep":
                    row1 += '<td class="gc-sep"></td>'
                    ci += 1
                elif dans_fenetre and ci == ci_debut and colspan_barre > 0:
                    import urllib.parse as _ul
                    proj_enc = _ul.quote(nom_projet)
                    tache_enc = _ul.quote(nom_tache)
                    row1 += (
                        f'<td colspan="{colspan_barre}" class="{cls_texte}" '
                        f'style="background:{couleur};">'
                        f'<a href="?gantt_projet={proj_enc}&gantt_tache={tache_enc}" '
                        f'target="_self" '
                        f'style="color:inherit;text-decoration:none;display:block;width:100%;height:100%;'
                        f'text-align:center;line-height:24px;">{nom_projet}</a>'
                        f'</td>'
                    )
                    ci += colspan_barre
                else:
                    est_aujourd_hui = colonnes[ci]["type"] == "jour" and colonnes[ci]["jour"] == today
                    cls = "gc-today" if est_aujourd_hui else "gc"
                    row1 += f'<td class="{cls}"></td>'
                    ci += 1
            row1 += '</tr>'

            # Ligne 2 : noms des ressources
            sep_cls = ' class="gsep"' if is_last else ''
            row2 = f'<tr{sep_cls}>'
            ci = 0
            while ci < nb_cols:
                col = colonnes[ci]
                if col["type"] == "sep":
                    row2 += '<td class="gc-sep2"></td>'
                    ci += 1
                elif dans_fenetre and ci == ci_debut and colspan_barre > 0:
                    row2 += (
                        f'<td colspan="{colspan_barre}" class="gb2" '
                        f'style="background:{couleur_sub};">{noms_str}</td>'
                    )
                    ci += colspan_barre
                else:
                    est_aujourd_hui = col["type"] == "jour" and col["jour"] == today
                    cls = "gc2-today" if est_aujourd_hui else "gc2"
                    row2 += f'<td class="{cls}"></td>'
                    ci += 1
            row2 += '</tr>'

            rows += row1 + row2

    # ── LEGENDE ──────────────────────────────────────────────────────────────
    types_presents = [t for t in ORDRE_TACHES if taches_par_type.get(t)]
    legende_html = '<div style="margin-top:12px;display:flex;flex-wrap:wrap;gap:12px;">'
    for t in types_presents:
        c = COULEURS_TACHES.get(t, "#A0A0A0")
        legende_html += (
            f'<span style="display:flex;align-items:center;gap:6px;font-size:0.85em;">'
            f'<span style="display:inline-block;width:16px;height:16px;'
            f'background:{c};border-radius:3px;"></span>{t}</span>'
        )
    legende_html += '</div>'

    html = f"""
    {css}
    <div class="gantt-wrap" id="{gantt_id}">
    <table class="gantt-table">
        {colgroup}
        <thead>{header}</thead>
        <tbody>{rows}</tbody>
    </table>
    </div>
    {legende_html}
    """
    return html


# -------------------------------------------------------
# ONGLET CALENDRIER
# -------------------------------------------------------

def gantt_ressources(ressources_base, data_proj, date_debut_grille, date_fin_grille, font_face=""):
    """
    Gantt par pôle/ressource.
    - Un bloc par pôle avec titre coloré
    - Une section par ressource dans le pôle
    - Barres colorées avec la couleur du pôle
    - Label : Nom projet — Sous-tâche
    """
    from donnees import POSTES, get_couleur_poste
    today = date.today()

    jours = get_jours(date_debut_grille, date_fin_grille)
    lundis = get_lundis(date_debut_grille, date_fin_grille)
    nb_jours = len(jours)

    colonnes = []
    for jour in jours:
        if jour.weekday() == 0 and jour != jours[0]:
            colonnes.append({"type": "sep"})
        colonnes.append({"type": "jour", "jour": jour})

    col_index = {col["jour"]: i for i, col in enumerate(colonnes) if col["type"] == "jour"}
    nb_cols = len(colonnes)
    nb_seps = sum(1 for c in colonnes if c["type"] == "sep")
    largeur_totale = nb_jours * 14 + nb_seps * 3

    # Construire index rapide : nom_ressource → liste de (nom_projet, sous_tache, debut, fin)
    ress_taches = {}
    for projet_nom, sous_taches_dict in data_proj.items():
        for nom_st, data_st in sous_taches_dict.items():
            for a in data_st.get("Assignations", []):
                nom_ress = a["Nom"]
                if nom_ress not in ress_taches:
                    ress_taches[nom_ress] = []
                # Chercher les dates dans Projets_gantt
                from donnees import get_couleur_projet
                import streamlit as _st
                for p in _st.session_state.Projets_gantt:
                    if p["projet"] == projet_nom:
                        for st_data in p["sous_taches"]:
                            if st_data["tache"] == nom_st:
                                ress_taches[nom_ress].append({
                                    "projet": projet_nom,
                                    "tache": nom_st,
                                    "start": st_data["start"],
                                    "end": st_data["end"],
                                })

    css = f"""
    <style>
    {font_face}
    body, table, td, th {{ font-family: 'GTWalsheim', sans-serif; }}
    .gantt-wrap {{ overflow-x: auto; width: 100%; }}
    .gantt-table {{
        border-collapse: separate;
        border-spacing: 0;
        width: {largeur_totale}px;
        table-layout: fixed;
        font-size: 0.82em;
        margin-bottom: 24px;
    }}
    .gh {{ text-align:left;padding:4px 6px;font-weight:bold;font-size:0.85em;
           background:#f5f5f5;border-bottom:2px solid #ccc;border-top:none;
           border-right:none;border-left:none;white-space:nowrap;overflow:hidden; }}
    .gh-sep {{ background:#999;width:3px;padding:0;border:none; }}
    .gc {{ padding:0;height:24px;border-right:1px solid #f0f0f0;
           border-top:none;border-bottom:none;border-left:none; }}
    .gc-today {{ padding:0;height:24px;border-right:1px solid #f0f0f0;
                 border-top:none;border-bottom:none;border-left:none;background:#FFFACD; }}
    .gc2 {{ padding:0;height:20px;border-right:1px solid #f0f0f0;
            border-top:none;border-bottom:none;border-left:none; }}
    .gc2-today {{ padding:0;height:20px;border-right:1px solid #f0f0f0;
                  border-top:none;border-bottom:none;border-left:none;background:#FFFACD; }}
    .gc-sep {{ background:#bbb;width:3px;padding:0;border:none;height:24px; }}
    .gc-sep2 {{ background:#bbb;width:3px;padding:0;border:none;height:20px; }}
    .gb1 {{ padding:0 8px;height:24px;vertical-align:middle;text-align:center;
            font-weight:bold;font-size:0.85em;color:#fff;white-space:nowrap;
            overflow:hidden;text-overflow:ellipsis;border:none; }}
    .gb1-dark {{ padding:0 8px;height:24px;vertical-align:middle;text-align:center;
                 font-weight:bold;font-size:0.85em;color:#222;white-space:nowrap;
                 overflow:hidden;text-overflow:ellipsis;border:none; }}
    .gb2 {{ padding:0 8px;height:20px;vertical-align:middle;font-size:0.78em;
            color:#333;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;border:none; }}
    .gsep td {{ border-bottom:2px solid #ccc !important; }}
    .ress-sep td {{ border-bottom:1px solid #e0e0e0 !important; }}
    </style>
    """

    colgroup = '<colgroup>'
    colgroup += '<col style="width:80px;min-width:80px;max-width:80px;">'  # colonne label nom
    for col in colonnes:
        if col["type"] == "sep":
            colgroup += '<col style="width:3px;min-width:3px;max-width:3px;">'
        else:
            colgroup += '<col style="width:14px;min-width:14px;">'
    colgroup += '</colgroup>'

    header = '<tr><th style="background:#f5f5f5;border-bottom:2px solid #ccc;"></th>'
    for lundi in lundis:
        jours_sem = [j for j in jours if lundi <= j < lundi + timedelta(weeks=1)]
        if not jours_sem:
            continue
        if lundi != lundis[0]:
            header += '<th class="gh-sep"></th>'
        label = f"Lun. {lundi.day}/{lundi.month}"
        header += f'<th colspan="{len(jours_sem)}" class="gh">{label}</th>'
    header += '</tr>'

    # Construire les blocs par pôle
    html_blocs = ""
    for poste, couleur_poste in POSTES.items():
        ress_du_poste = [r for r in ressources_base if r.get("Poste") == poste]
        if not ress_du_poste:
            continue

        # Titre du pôle
        html_blocs += (
            f'<div style="margin-top:20px;margin-bottom:6px;font-weight:bold;'
            f'font-size:1em;color:{couleur_poste};border-left:4px solid {couleur_poste};'
            f'padding-left:8px;">{poste}</div>'
        )

        # Couleur texte selon luminosité
        r_int = int(couleur_poste.lstrip("#")[0:2], 16)
        g_int = int(couleur_poste.lstrip("#")[2:4], 16)
        b_int = int(couleur_poste.lstrip("#")[4:6], 16)
        luminosite = 0.299*r_int + 0.587*g_int + 0.114*b_int
        cls_texte = "gb1-dark" if luminosite > 160 else "gb1"
        couleur_sub = couleur_poste + "99"

        rows = ""
        for r_idx, ress in enumerate(ress_du_poste):
            nom_ress = ress["Nom"]
            taches_ress = ress_taches.get(nom_ress, [])
            is_last_ress = r_idx == len(ress_du_poste) - 1

            if not taches_ress:
                # Ressource sans assignation : 2 lignes vides
                row1 = f'<tr><td style="padding:2px 8px;font-size:0.82em;font-weight:bold;" colspan="1">{nom_ress}</td>'
                for col in colonnes:
                    if col["type"] == "sep":
                        row1 += '<td class="gc-sep"></td>'
                    elif col["jour"] == today:
                        row1 += '<td class="gc-today"></td>'
                    else:
                        row1 += '<td class="gc"></td>'
                row1 += '</tr>'
                sep_cls = ' class="gsep"' if is_last_ress else ' class="ress-sep"'
                row2 = f'<tr{sep_cls}><td></td>'
                for col in colonnes:
                    if col["type"] == "sep":
                        row2 += '<td class="gc-sep2"></td>'
                    elif col["jour"] == today:
                        row2 += '<td class="gc2-today"></td>'
                    else:
                        row2 += '<td class="gc2"></td>'
                row2 += '</tr>'
                rows += row1 + row2
            else:
                for t_idx, tache in enumerate(taches_ress):
                    is_last_tache = t_idx == len(taches_ress) - 1
                    debut_st = date.fromisoformat(tache["start"])
                    fin_st = date.fromisoformat(tache["end"])
                    label_barre = f"{tache['projet']} — {tache['tache']}"

                    debut_visible = max(debut_st, date_debut_grille)
                    fin_visible = min(fin_st, date_fin_grille)
                    dans_fenetre = debut_visible < fin_visible

                    if dans_fenetre:
                        ci_debut = col_index.get(debut_visible, 0)
                        ci_fin = col_index.get(fin_visible, nb_cols)
                        colspan_barre = ci_fin - ci_debut
                    else:
                        ci_debut = None
                        colspan_barre = 0

                    # Label ressource uniquement sur la première tâche
                    label_ress = nom_ress if t_idx == 0 else ""

                    # Ligne 1
                    row1 = f'<tr><td style="padding:2px 8px;font-size:0.82em;font-weight:bold;white-space:nowrap;">{label_ress}</td>'
                    ci = 0
                    while ci < nb_cols:
                        col = colonnes[ci]
                        if col["type"] == "sep":
                            row1 += '<td class="gc-sep"></td>'
                            ci += 1
                        elif dans_fenetre and ci == ci_debut and colspan_barre > 0:
                            row1 += (
                                f'<td colspan="{colspan_barre}" class="{cls_texte}" '
                                f'style="background:{couleur_poste};">{label_barre}</td>'
                            )
                            ci += colspan_barre
                        else:
                            est_today = col["type"] == "jour" and col["jour"] == today
                            row1 += f'<td class="{"gc-today" if est_today else "gc"}"></td>'
                            ci += 1
                    row1 += '</tr>'

                    # Ligne 2 vide (cohérence visuelle avec vue projets)
                    sep = is_last_tache and is_last_ress
                    sep_cls = ' class="gsep"' if sep else (' class="ress-sep"' if is_last_tache else '')
                    row2 = f'<tr{sep_cls}><td></td>'
                    ci = 0
                    while ci < nb_cols:
                        col = colonnes[ci]
                        if col["type"] == "sep":
                            row2 += '<td class="gc-sep2"></td>'
                            ci += 1
                        elif dans_fenetre and ci == ci_debut and colspan_barre > 0:
                            row2 += f'<td colspan="{colspan_barre}" class="gb2" style="background:{couleur_sub};"></td>'
                            ci += colspan_barre
                        else:
                            est_today = col["type"] == "jour" and col["jour"] == today
                            row2 += f'<td class="{"gc2-today" if est_today else "gc2"}"></td>'
                            ci += 1
                    row2 += '</tr>'
                    rows += row1 + row2

        html_blocs += f"""
        <div class="gantt-wrap">
        <table class="gantt-table">
            {colgroup}
            <thead>{header}</thead>
            <tbody>{rows}</tbody>
        </table>
        </div>
        """

    return f"{css}{html_blocs}"


# -------------------------------------------------------
# ONGLET CALENDRIER
# -------------------------------------------------------

def calendrier_tab():
    st.subheader('Calendrier')
    selection = st.pills(
        " ",
        ["Par projet", "Par ressource", "Absences"],
        selection_mode="single",
        default="Par projet"
    )

    if selection in ("Par projet", "Par ressource"):
        projets = st.session_state.Projets_gantt
        data_proj = st.session_state.Data_proj

        # Charger la police GT Walsheim
        import base64 as _b64, os as _os
        font_face = ""
        font_path = _os.path.join(_os.path.dirname(__file__), "fonts", "GT-Walsheim-Regular.ttf")
        try:
            with open(font_path, "rb") as f:
                font_b64 = _b64.b64encode(f.read()).decode("utf-8")
            font_face = f"""
            @font-face {{
                font-family: 'GTWalsheim';
                src: url(data:font/truetype;base64,{font_b64}) format('truetype');
            }}
            """
        except FileNotFoundError:
            pass

        # Plage de dates commune
        toutes_dates = [
            date.fromisoformat(st_data[cle])
            for p in projets
            for st_data in p.get("sous_taches", [])
            for cle in ("start", "end")
        ]
        today = date.today()
        if toutes_dates:
            date_debut_grille = min(min(toutes_dates), today) - timedelta(days=7)
            date_fin_grille = max(toutes_dates) + timedelta(days=7)
        else:
            date_debut_grille = today - timedelta(days=7)
            date_fin_grille = today + timedelta(weeks=12)

    if selection == "Par projet":
        st.markdown("#### Vue d'ensemble")
        if projets:
            html_global = gantt_tableau(projets, data_proj, date_debut_grille, date_fin_grille, "gantt_global", font_face)
            st.markdown(html_global, unsafe_allow_html=True)

            col_csv, col_xl, _ = st.columns([1, 1, 6])
            with col_csv:
                st.download_button(
                    "⬇ CSV",
                    data=_build_export_df(projets, data_proj).to_csv(index=False, encoding="utf-8-sig"),
                    file_name="gantt_global.csv",
                    mime="text/csv",
                    key="dl_csv_global"
                )
            with col_xl:
                st.download_button(
                    "⬇ Excel",
                    data=_to_excel_gantt(projets, data_proj),
                    file_name="gantt_global.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_xl_global"
                )
        else:
            st.info("Aucun projet à afficher.")

        params = st.query_params
        if "gantt_projet" in params and "gantt_tache" in params:
            nom_projet_clic = params["gantt_projet"]
            nom_tache_clic = params["gantt_tache"]
            st.query_params.clear()
            st.session_state["choix_projet"] = nom_projet_clic
            st.session_state["choix_sous_tache"] = nom_tache_clic
            st.switch_page("pages/04_Assignation.py")

        st.markdown("#### Vue détaillée")
        if projets:
            noms_projets = [p["projet"] for p in projets]
            projet_choisi = st.selectbox(
                "Sélectionner un projet :",
                options=noms_projets,
                key="gantt_detail_projet"
            )
            projet_data = next(p for p in projets if p["projet"] == projet_choisi)
            if projet_data["sous_taches"]:
                dates_proj = [
                    date.fromisoformat(st_data[cle])
                    for st_data in projet_data["sous_taches"]
                    for cle in ("start", "end")
                ]
                d_debut = min(min(dates_proj), today) - timedelta(days=7)
                d_fin = max(dates_proj) + timedelta(days=7)
                html_detail = gantt_tableau([projet_data], data_proj, d_debut, d_fin, "gantt_detail", font_face)
                st.markdown(html_detail, unsafe_allow_html=True)

                col_csv2, col_xl2, _ = st.columns([1, 1, 6])
                with col_csv2:
                    st.download_button(
                        "⬇ CSV",
                        data=_build_export_df([projet_data], data_proj).to_csv(index=False, encoding="utf-8-sig"),
                        file_name=f"gantt_{projet_choisi}.csv",
                        mime="text/csv",
                        key="dl_csv_detail"
                    )
                with col_xl2:
                    st.download_button(
                        "⬇ Excel",
                        data=_to_excel_gantt([projet_data], data_proj),
                        file_name=f"gantt_{projet_choisi}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_xl_detail"
                    )
            else:
                st.info("Ce projet n'a pas encore de sous-tâches.")

    if selection == "Par ressource":
        ressources = st.session_state.Ressources_base
        if ressources:
            html_ress = gantt_ressources(ressources, data_proj, date_debut_grille, date_fin_grille, font_face)
            st.markdown(html_ress, unsafe_allow_html=True)
        else:
            st.info("Aucune ressource créée.")

    if selection == "Absences":
        if st.session_state.get("msg_succes"):
            st.success(st.session_state.msg_succes)
            st.session_state.msg_succes = None

        noms_ressources = [r["Nom"] for r in st.session_state.Ressources_base]

        # Construire la liste de toutes les absences avec leur index
        # Format : [(nom_ress, idx_absence, label_affichage)]
        toutes_absences = []
        for r in st.session_state.Ressources_base:
            for idx, absence in enumerate(r.get("absences", [])):
                label = f"{r['Nom']} — {fmt_date(absence['start'])} → {fmt_date(absence['end'])}"
                toutes_absences.append((r["Nom"], idx, label, absence))

        # ── Ajout d'une absence ──
        st.markdown("#### Ajouter une absence")
        with st.form("form_absence", clear_on_submit=True):
            col_nom, col_debut, col_fin = st.columns([2, 1, 1])
            with col_nom:
                nom_choisi = st.selectbox("Ressource", options=noms_ressources)
            with col_debut:
                date_debut = st.date_input("Début", value=date.today())
            with col_fin:
                date_fin = st.date_input("Fin", value=date.today() + timedelta(days=7))
            submitted = st.form_submit_button("➕ Ajouter")

        if submitted:
            if date_fin <= date_debut:
                st.error("La date de fin doit être après la date de début.")
            else:
                for r in st.session_state.Ressources_base:
                    if r["Nom"] == nom_choisi:
                        if "absences" not in r:
                            r["absences"] = []
                        r["absences"].append({
                            "start": date_debut.isoformat(),
                            "end": date_fin.isoformat()
                        })
                        break
                nouveau_sha = sauvegarder_ressources_github(
                    st.session_state.Ressources_base,
                    st.session_state.ressources_sha
                )
                st.session_state.ressources_sha = nouveau_sha
                st.session_state.msg_succes = f"Absence de {nom_choisi} ajoutée."
                st.rerun()

        # ── Modifier ou supprimer une absence ──
        if toutes_absences:
            st.markdown("#### Modifier ou supprimer une absence")
            options_labels = [a[2] for a in toutes_absences]
            choix_label = st.selectbox(
                "Sélectionner une absence",
                options=options_labels,
                key="absence_selectbox"
            )
            idx_choix = options_labels.index(choix_label)
            nom_ress_sel, idx_abs_sel, _, absence_sel = toutes_absences[idx_choix]

            with st.form("form_modif_absence"):
                col_nom2, col_debut2, col_fin2 = st.columns([2, 1, 1])
                with col_nom2:
                    nouveau_nom = st.selectbox(
                        "Ressource",
                        options=noms_ressources,
                        index=noms_ressources.index(nom_ress_sel)
                    )
                with col_debut2:
                    nouveau_debut = st.date_input(
                        "Début",
                        value=date.fromisoformat(absence_sel["start"])
                    )
                with col_fin2:
                    nouveau_fin = st.date_input(
                        "Fin",
                        value=date.fromisoformat(absence_sel["end"])
                    )
                col_save, col_del = st.columns([1, 1])
                with col_save:
                    sauvegarder = st.form_submit_button("✅ Modifier")
                with col_del:
                    supprimer = st.form_submit_button("🗑 Supprimer")

            if sauvegarder:
                if nouveau_fin <= nouveau_debut:
                    st.error("La date de fin doit être après la date de début.")
                else:
                    # Supprimer l'ancienne absence
                    for r in st.session_state.Ressources_base:
                        if r["Nom"] == nom_ress_sel:
                            r["absences"].pop(idx_abs_sel)
                            break
                    # Ajouter la nouvelle (potentiellement sur une autre ressource)
                    for r in st.session_state.Ressources_base:
                        if r["Nom"] == nouveau_nom:
                            if "absences" not in r:
                                r["absences"] = []
                            r["absences"].append({
                                "start": nouveau_debut.isoformat(),
                                "end": nouveau_fin.isoformat()
                            })
                            break
                    nouveau_sha = sauvegarder_ressources_github(
                        st.session_state.Ressources_base,
                        st.session_state.ressources_sha
                    )
                    st.session_state.ressources_sha = nouveau_sha
                    st.session_state.msg_succes = "Absence modifiée."
                    st.rerun()

            if supprimer:
                for r in st.session_state.Ressources_base:
                    if r["Nom"] == nom_ress_sel:
                        r["absences"].pop(idx_abs_sel)
                        break
                nouveau_sha = sauvegarder_ressources_github(
                    st.session_state.Ressources_base,
                    st.session_state.ressources_sha
                )
                st.session_state.ressources_sha = nouveau_sha
                st.session_state.msg_succes = f"Absence de {nom_ress_sel} supprimée."
                st.rerun()

        absences_cal = build_absences_cal()
        calendar(events=absences_cal, options=Options_cal)
